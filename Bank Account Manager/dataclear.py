import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

sheet.title = 'name'
sheet.cell(row = 1,column = 1).value = '1'
sheet.cell(row = 1,column = 2).value = 'NAME'
sheet.cell(row = 1,column = 3).value = 'PIN'

wb.save('name.xlsx')

wb = openpyxl.Workbook()

sheet = wb.active

sheet.title = 'ac'
sheet.cell(row = 1,column = 1).value = '100'
sheet.cell(row = 1,column = 2).value = 'AC NUMBER'
sheet.cell(row = 1,column = 3).value = 'AC TYPE'
sheet.cell(row = 1,column = 4).value = 'BALACE'

wb.save('ac.xlsx')
