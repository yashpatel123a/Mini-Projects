from account import *
import openpyxl,os

def find_in_xl(sheet, name):
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(i, 2).value == name:
            return i
    return -1


def open_account(customer, index):
    clean()
    #load file name.xlsx and ac.xlsx 
    rd1 = openpyxl.load_workbook('name.xlsx') 
    write_name = rd1.active   
    rd2 = openpyxl.load_workbook('ac.xlsx')
    write_ac = rd2.active

    if customer.total_account() > 5:
        print('Maximum account per customer.')
        delay()

    else:
        gen_ac = int(write_ac.cell(row = 1, column = 1).value)#loading account number generater located at (1,1) in ac.xlsx 

        print('----------------------------------------')
        print('1 - open checking account')
        print('2 - open savings account')
        print('3 - open business account')
        print('0 - back')
        print('----------------------------------------')
        c4 = input('Enter your choice: ')

        if c4 != '0':
            depo = float(input('Enter your opening deposit: '))
            write_name.cell(row = index,column = 1).value = str(customer.total_account() + 1)#update account per user in name.xlsx

        if c4 == '1':
            customer.open_checking(gen_ac,depo)#open checing account
            write_ac.cell(row = gen_ac - 98,column = 2).value = str(gen_ac)#adding account number,account type,and updating account number generator(1,1) in ac.xlsx
            write_ac.cell(row = gen_ac - 98,column = 3).value = 'C'
            write_ac.cell(row = gen_ac - 98,column = 4).value = str(depo)
            gen_ac +=1
            write_ac.cell(row = 1,column = 1).value = str(gen_ac)

        if c4 == '2':
            customer.open_savings(gen_ac,depo)
            write_ac.cell(row = gen_ac - 98,column = 2).value = str(gen_ac)#same as above with savings account
            write_ac.cell(row = gen_ac - 98,column = 3).value = 'S'
            write_ac.cell(row = gen_ac - 98,column = 4).value = str(depo)
            gen_ac +=1
            write_ac.cell(row = 1,column = 1).value = str(gen_ac)

        if c4 == '3':
            customer.open_business(gen_ac,depo)
            write_ac.cell(row = gen_ac - 98,column = 2).value = str(gen_ac)#same as above with business account
            write_ac.cell(row = gen_ac - 98,column = 3).value = 'B'
            write_ac.cell(row = gen_ac - 98,column = 4).value = str(depo)
            gen_ac +=1
            write_ac.cell(row = 1,column = 1).value = str(gen_ac)

        if c4 == '0':
            pass#do nothing

        write_name.cell(row = index,column = customer.total_account()+3).value = str(gen_ac-1)#add account number at name.xlsx in customer's row
        rd1.save('name.xlsx')
        rd2.save('ac.xlsx')

def managment(customer,index):
    clean()

    print(f'Welcome {customer.name}')
    while True:
        clean()

        print('----------------------------------------')
        print('1 - Open new Account')
        print('2 - deposit amount')
        print('3 - withdraw amount')
        print('4 - List of account and total balance')
        print('0 - Logout')
        print('----------------------------------------')
        c3 = input('Enter your choice: ')

        if c3 == '1':
            open_account(customer,index)

        if c3 == '2':

            depo = float(input('Enter a amount of deposit: '))#taking deposit amount,account type,account number
            ac_type = input('Enter account type: ')
            ac_number = input('Enter account number: ')
            
            rd2 = openpyxl.load_workbook('ac.xlsx')
            write_ac = rd2.active

            rowa = find_in_xl(write_ac,ac_number)#return a row number in ac.xlsx where that account number located

            final_bal = float(write_ac.cell(row = rowa,column = 4).value) + depo#loading balance and update it

            flag = make_dep(customer,ac_type[0].upper(),int(ac_number),depo)
            if flag == 0:
                print('deposit successfull')
                write_ac.cell(row = rowa,column = 4).value = float(final_bal)#updating balance in ac.xlsx
                rd2.save('ac.xlsx')
            else:
                print('deposit unsuccessfull')
            delay()
            

        if c3 == '3':

            withdraw = float(input('Enetr a amount of withdraw: '))#taking deposit amount,account type,account number
            ac_type = input('Enter account type: ')
            ac_number = input('Enter account number: ')

            clean()
            ac_pin = input('Enter your pin: ')

            if ac_pin == customer.PIN:
                flag = make_wd(customer,ac_type[0].upper(),int(ac_number),withdraw)#flag = 0 for successfull withdraw

                if flag == 0:

                    rd2 = openpyxl.load_workbook('ac.xlsx')
                    write_ac = rd2.active

                    rowa = find_in_xl(write_ac,ac_number)#return a row number in ac.xlsx where that account number located

                    final_bal = float(write_ac.cell(row = rowa,column = 4).value) - withdraw

                    write_ac.cell(row = rowa,column = 4).value = float(final_bal)#updating balance in ac.xlsx
                    rd2.save('ac.xlsx')

                    print('withdraw successfull')

                else:
                    print('withdraw unsuccessfull')
                delay()
            else:

                print('INVALID PIN')
                delay()  

        if c3 == '4':
            clean()
            customer.get_total_deposits()
            delay()

        if c3 == '0':
            break
            
def clean():
        try:
	        os.system('cls')
        except:
            os.system('clear')

def delay():
    temp = input('\nPress any key to continue.')


def sign_up():

    clean()

    wb1 = openpyxl.load_workbook('name.xlsx')
    write_name = wb1.active

    cus_number = int(write_name.cell(1,1).value)#loading total number of customers to calculate where to add next data

    name = input('Enter your name: ')
    pin = input('Set your password: ')

    write_name.cell(row = cus_number + 1,column = 2).value = name#writing name,PIN,customer's total account and update customer number
    write_name.cell(row = cus_number + 1,column = 3).value = pin
    write_name.cell(row = cus_number + 1,column = 1).value = '0'
    write_name.cell(row = 1,column = 1).value = str(cus_number+1)

    wb1.save('name.xlsx')

def sign_in():

    clean()

    rd1 = openpyxl.load_workbook('name.xlsx')
    write_name = rd1.active   
    rd2 = openpyxl.load_workbook('ac.xlsx')
    write_ac = rd2.active   

    si_name = input('Enter your name to sign in:')

    rown = find_in_xl(write_name,si_name)#find the name in name.xlsx and return row number if not found than return -1

    if rown != -1:

        sheet_pin = write_name.cell(rown,3).value#loading True PIN number of customer

        clean()
        si_pin = input('Enter your pin: ')

        if si_pin == sheet_pin:

            customer = Customer(si_name,si_pin)#Create customer named object with type Customer
    
            number_of_ac = write_name.cell(rown,1).value#load the number of account that customer holds

            for i in range(int(number_of_ac)):

                acn = write_name.cell(rown,i+4).value#take account number

                rowa = find_in_xl(write_ac,acn)#find account number in ac.xlsx and return row number

                bal = float(write_ac.cell(rowa,4).value)#load balance from that row number

                ty = write_ac.cell(rowa,3).value#load type of the account 
                #adding accounts in customer object
                if ty == 'C':
                    customer.open_checking(int(acn),bal)
                if ty == 'S':
                    customer.open_savings(int(acn),bal)
                if ty == 'B':
                    customer.open_business(int(acn),bal)
            
            managment(customer,rown)

        else:
            print('INVALID PIN')
    else:

        print("INVALID USERNAME")
    delay()

while True:
    clean()

    print('---------------------------------------')
    print('       Welcome to your own bank')
    print('---------------------------------------')

    print('1 - sign up')
    print('2 - sign in')
    print('0 - Exit')
    c1 = input('Enter your choice: ')

    if c1 == '1':
        sign_up()

    if c1 == '2':
        sign_in()
        
    if c1 == '3':
        passw = input('Enter password:')
        if passw == 'd4dog':
            wb=openpyxl.load_workbook('name.xlsx')
            sheet=wb.active
            max_row=sheet.max_row
            max_column=sheet.max_column
            for i in range(1,max_row+1):
                for j in range(1,max_column+1):            
                    cell_obj=sheet.cell(row=i,column=j)        
                    print(cell_obj.value,end=' | ') 
                print('\n')
            print('--------------------------------------------------------')
            wb=openpyxl.load_workbook('ac.xlsx')
            sheet=wb.active
            max_row=sheet.max_row
            max_column=sheet.max_column
            for i in range(1,max_row+1):
                for j in range(1,max_column+1):            
                    cell_obj=sheet.cell(row=i,column=j)        
                    print(cell_obj.value,end=' | ') 
                print('\n')
            delay()

    if c1 == '0':
        clean()
        break
    

